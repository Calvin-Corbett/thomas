"""File content extractors for chat document uploads.

Extracts readable text from uploaded files so Thomas can understand
documents the user shares in chat. Supports PDF, DOCX, and XLSX.

Feature parity gap: Claude/ChatGPT/Grok all read uploaded PDFs and
Office docs natively. Thomas previously only handled plain text.

Architecture context for AI agents:
  - Called from chat_aiohttp_part02.py during document attachment processing
  - Returns plain text suitable for including in the LLM prompt
  - Dependencies: pypdf2, python-docx, openpyxl (install via pip)
  - Falls back gracefully if dependencies aren't installed

Usage:
    text = extract_file_text("/path/to/file.pdf")
    if text:
        # Include in prompt context
"""

from __future__ import annotations

import logging
from pathlib import Path

log = logging.getLogger(__name__)

# Max characters to extract from any single file
_MAX_EXTRACT_CHARS = 100_000


def extract_file_text(path: str | Path, *, max_chars: int = _MAX_EXTRACT_CHARS) -> str | None:
    """Extract readable text from a file.

    Auto-detects file type from extension. Returns None if the file type
    is unsupported or extraction fails.

    Supported formats:
      - .pdf (via pypdf)
      - .docx (via python-docx)
      - .xlsx / .csv (via openpyxl / stdlib csv)
      - .txt / .md / .json / .py / etc. (plain text)
    """
    p = Path(path)
    if not p.exists():
        return None

    suffix = p.suffix.lower()

    try:
        if suffix == ".pdf":
            return _extract_pdf(p, max_chars)
        elif suffix == ".docx":
            return _extract_docx(p, max_chars)
        elif suffix in (".xlsx", ".xls"):
            return _extract_xlsx(p, max_chars)
        elif suffix == ".csv":
            return _extract_csv(p, max_chars)
        elif suffix in (
            ".txt",
            ".md",
            ".json",
            ".py",
            ".js",
            ".ts",
            ".html",
            ".css",
            ".yaml",
            ".yml",
            ".toml",
            ".xml",
            ".log",
            ".sh",
            ".bat",
            ".cmd",
            ".ps1",
            ".rs",
            ".go",
            ".java",
            ".c",
            ".cpp",
            ".h",
            ".rb",
            ".php",
            ".sql",
            ".r",
            ".swift",
            ".kt",
            ".lua",
            ".pl",
            ".env",
            ".ini",
            ".cfg",
        ):
            return _extract_text(p, max_chars)
        else:
            # Try as plain text
            return _extract_text(p, max_chars)
    except Exception as exc:
        log.warning("Failed to extract text from %s: %s", path, exc)
        return None


def _extract_pdf(path: Path, max_chars: int) -> str | None:
    """Extract text from a PDF file."""
    try:
        from pypdf import PdfReader
    except ImportError:
        try:
            from PyPDF2 import PdfReader  # type: ignore[no-redef]
        except ImportError:
            log.debug("PDF extraction unavailable — install pypdf: pip install pypdf")
            return None

    reader = PdfReader(str(path))
    pages: list[str] = []
    total = 0
    for page in reader.pages:
        text = page.extract_text() or ""
        if total + len(text) > max_chars:
            text = text[: max_chars - total]
            pages.append(text)
            break
        pages.append(text)
        total += len(text)

    result = "\n\n".join(pages).strip()
    return result if result else None


def _extract_docx(path: Path, max_chars: int) -> str | None:
    """Extract text from a Word document."""
    try:
        from docx import Document  # type: ignore[import-untyped]
    except ImportError:
        log.debug("DOCX extraction unavailable — install python-docx: pip install python-docx")
        return None

    doc = Document(str(path))
    paragraphs: list[str] = []
    total = 0
    for para in doc.paragraphs:
        text = para.text or ""
        if total + len(text) > max_chars:
            text = text[: max_chars - total]
            paragraphs.append(text)
            break
        paragraphs.append(text)
        total += len(text)

    result = "\n".join(paragraphs).strip()
    return result if result else None


def _extract_xlsx(path: Path, max_chars: int) -> str | None:
    """Extract text from an Excel spreadsheet."""
    try:
        from openpyxl import load_workbook  # type: ignore[import-untyped]
    except ImportError:
        log.debug("XLSX extraction unavailable — install openpyxl: pip install openpyxl")
        return None

    wb = load_workbook(str(path), read_only=True, data_only=True)
    sheets: list[str] = []
    total = 0

    for ws in wb.worksheets:
        rows: list[str] = []
        rows.append(f"--- Sheet: {ws.title} ---")
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            row_text = "\t".join(cells)
            if total + len(row_text) > max_chars:
                break
            rows.append(row_text)
            total += len(row_text) + 1
        sheets.append("\n".join(rows))
        if total >= max_chars:
            break

    wb.close()
    result = "\n\n".join(sheets).strip()
    return result if result else None


def _extract_csv(path: Path, max_chars: int) -> str | None:
    """Extract text from a CSV file."""
    import csv

    rows: list[str] = []
    total = 0
    with open(path, newline="", encoding="utf-8", errors="replace") as f:
        reader = csv.reader(f)
        for row in reader:
            row_text = "\t".join(row)
            if total + len(row_text) > max_chars:
                break
            rows.append(row_text)
            total += len(row_text) + 1
    result = "\n".join(rows).strip()
    return result if result else None


def _extract_text(path: Path, max_chars: int) -> str | None:
    """Read a plain text file."""
    try:
        text = path.read_text(encoding="utf-8", errors="replace")
        if len(text) > max_chars:
            text = text[:max_chars] + "\n... (truncated)"
        return text.strip() if text.strip() else None
    except Exception:
        return None


def detect_and_extract(
    file_path: str | Path,
    file_name: str = "",
) -> tuple[str, str | None]:
    """Detect file type and extract text.

    Returns (display_name, extracted_text).
    If extraction fails, returns (display_name, None).
    """
    p = Path(file_path)
    name = file_name or p.name
    text = extract_file_text(p)
    return name, text
